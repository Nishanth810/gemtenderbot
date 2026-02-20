# =============================================================================
#  GeM BidPlus Tender Bot — ULTIMATE FINAL VERSION
#
#  ✅ Uses friend's working URL pattern: showbidDocument/<numeric_id>
#  ✅ Organization fallback from target (not from API)
#  ✅ Links work perfectly
#  ✅ Excel shows correct org names
#  ✅ Email to multiple receivers (STARTTLS + SSL fallback)
#  ✅ Windows sleep prevention
#  ✅ Duplicates shown in RED on separate sheet
#  ✅ Rich HTML email format
#  ✅ RA NO bids (Rate Agreement) filtered out completely
#
#  INSTALL:  pip install requests pandas openpyxl python-dotenv
#  RUN:      python gem_scraper_ULTIMATE.py
# =============================================================================

import os, re, sys, json, time, logging, smtplib, traceback
import requests
from datetime import datetime
from pathlib import Path
from logging.handlers import RotatingFileHandler
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import Dict, List, Tuple

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
#  CONFIG
# =============================================================================
BASE_URL    = "https://bidplus.gem.gov.in"
SEARCH_URL  = BASE_URL + "/search-bids"
ADVANCE_URL = BASE_URL + "/advance-search"

DELAY_PAGE    = 2
DELAY_ORG     = 3
DELAY_REFRESH = 5
MAX_PAGES     = 2000
PAGE_SIZE     = 10

EXCEL_FOLDER  = "Excel_Reports"
LOG_FOLDER    = "logs"
HISTORY_FILE  = "gem_tender_history.json"

COLOR_HEADER   = PatternFill("solid", fgColor="1F4E79")
COLOR_NEW      = PatternFill("solid", fgColor="C6EFCE")
COLOR_CHANGED  = PatternFill("solid", fgColor="FFEB9C")
COLOR_EXISTING = PatternFill("solid", fgColor="EBF3FB")
COLOR_ALT      = PatternFill("solid", fgColor="D9E1F2")
COLOR_DUPE     = PatternFill("solid", fgColor="FF9999")   # RED for duplicates
FONT_HEADER    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
FONT_LINK      = Font(name="Arial", color="0563C1", underline="single", size=10)
FONT_DATA      = Font(name="Arial", size=10)
FONT_DUPE      = Font(name="Arial", size=10, color="CC0000", bold=True)
THIN           = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# =============================================================================
#  TARGETS
# =============================================================================
TARGETS = [
    {"ministry": "Department of Space", "organization": "indian space research organization"},
    {"ministry": "Ministry of Defence", "organization": "Bharat Dynamics Limited"},
    {"ministry": "Ministry of Defence", "organization": "Bharat Earth Movers Limited (BEML)"},
    {"ministry": "Ministry of Defence", "organization": "Bharat Electronics Limited (BEL)"},
    {"ministry": "Ministry of Defence", "organization": "Defence Research and Development Organisation (DRDO)"},
    {"ministry": "Ministry of Defence", "organization": "Hindustan Aeronautics Limited (HAL)"},
    {"ministry": "Ministry of Defence", "organization": "Hindustan Shipyard Limited (HSL)"},
    {"ministry": "Ministry of Defence", "organization": "HQIDS"},
    {"ministry": "Ministry of Defence", "organization": "Indian Air Force"},
    {"ministry": "Ministry of Defence", "organization": "Indian Army"},
    {"ministry": "Ministry of Defence", "organization": "Indian Navy"},
    {"ministry": "Ministry of Defence", "organization": "Office of DG (ECS)"},
    {"ministry": "Ministry of Defence", "organization": "Office of DG (HR)"},
    {"ministry": "Ministry of Defence", "organization": "Office of DG (LS)"},
    {"ministry": "Ministry of Defence", "organization": "Office of DG (ACE)"},
    {"ministry": "Ministry of Defence", "organization": "Office of DG (MSS)"},
    {"ministry": "Ministry of Defence", "organization": "Office of DG R & D (R & M)"},
    {"ministry": "Ministry of Defence", "organization": "Office of DG R & D (SAM)"},
    {"ministry": "Ministry of Defence", "organization": "Office of DG ( MED & COS)"},
    {"ministry": "Ministry of Defence", "organization": "Office of DG ( NS & M)"},
    {"ministry": "Ministry of Defence", "organization": "Office of DG (Aero)"},
    {"ministry": "Ministry of Defence", "organization": "Yantra India Limited"},
    {"ministry": "PMO", "organization": "ANUSHAKTI VIDHYUT NIGAM Limited"},
    {"ministry": "PMO", "organization": "Atomic Energy Education Society (AEES)"},
    {"ministry": "PMO", "organization": "Atomic Energy Regulatory Board"},
    {"ministry": "PMO", "organization": "ATOMIC MINERALS DIRECTORATE"},
    {"ministry": "PMO", "organization": "Bhabha Atomic Research Centre"},
    {"ministry": "PMO", "organization": "BHARATIYA NABHIKIYA VIDYUT NIGAM Limited"},
    {"ministry": "PMO", "organization": "Board of Radiation and Isotope Technology"},
    {"ministry": "PMO", "organization": "Centre for Excellence in Basic Sciences"},
    {"ministry": "PMO", "organization": "Directorate of Purchase and Stores"},
    {"ministry": "PMO", "organization": "Dr Bhubanesar Borooah Cancer Institute BBCI"},
    {"ministry": "PMO", "organization": "ELECTRONICS Corporation OF INDIA Limited"},
    {"ministry": "PMO", "organization": "HEAVY WATER BOARD MUMBAI"},
    {"ministry": "PMO", "organization": "Homi Bhabha Cancer Hospital Sangrur"},
    {"ministry": "PMO", "organization": "Homi Bhabha National Institute"},
    {"ministry": "PMO", "organization": "Indian Institute of Space Science Technology"},
    {"ministry": "PMO", "organization": "Institute for Plasma Research"},
    {"ministry": "PMO", "organization": "IREL (India) Limited"},
    {"ministry": "PMO", "organization": "IREL India Limited"},
    {"ministry": "PMO", "organization": "ISRO"},
    {"ministry": "PMO", "organization": "ITER India IPR"},
    {"ministry": "PMO", "organization": "National Security Advisory Board (NSAB)"},
    {"ministry": "PMO", "organization": "NTRO"},
    {"ministry": "PMO", "organization": "NUCLEAR POWER CORPORATION OF INDIA LIMITED"},
    {"ministry": "PMO", "organization": "Saha Institute of Nuclear Physics"},
    {"ministry": "PMO", "organization": "Semi Conductor Laboratory"},
    {"ministry": "PMO", "organization": "TATA INSTITUTE OF FUNDAMENTAL RESEARCH HYDERABAD"},
    {"ministry": "PMO", "organization": "Tata Institute of Fundamental Research MUMBAI"},
    {"ministry": "PMO", "organization": "Tata Memorial Center \u2013 Advanced Center for Treatment, Research and Education in Cancer (ACTREC)"},
    {"ministry": "PMO", "organization": "Tata Memorial Centre"},
    {"ministry": "PMO", "organization": "Tata Memorial Centre  HBCH RC VIZAG"},
    {"ministry": "PMO", "organization": "TATA MEMORIAL CENTRE MPMMCC AND HBCH VARANASI Madh"},
    {"ministry": "PMO", "organization": "TIFR BALLOON FACILITY HYDERABAD"},
    {"ministry": "PMO", "organization": "TIFR CENTRE FOR APPLICABLE MATHEMATICS"},
    {"ministry": "PMO", "organization": "URANIUM CORPORATION OF INDIA LIMITED"},
    {"ministry": "PMO", "organization": "VARIABLE ENERGY CYCLOTRON CENTRE"},
]

# =============================================================================
#  LOGGING
# =============================================================================
def setup_logger() -> logging.Logger:
    os.makedirs(LOG_FOLDER, exist_ok=True)
    logger = logging.getLogger("gem_bot")
    if logger.handlers:
        return logger
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s", "%Y-%m-%d %H:%M:%S")
    ch = logging.StreamHandler()
    ch.setFormatter(fmt)
    logger.addHandler(ch)
    fh = RotatingFileHandler(
        os.path.join(LOG_FOLDER, f"gem_bot_{datetime.now().strftime('%Y%m%d')}.log"),
        maxBytes=10*1024*1024, backupCount=5, encoding="utf-8",
    )
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger

LOG = setup_logger()

# Always load .env from the app directory and override any stale session vars.
load_dotenv(dotenv_path=Path(__file__).resolve().parent / ".env", override=True)
SENDER_EMAIL    = os.getenv("SENDER_EMAIL", "")
APP_PASSWORD    = os.getenv("APP_PASSWORD", "")
RECEIVER_EMAILS = [e.strip() for e in os.getenv("RECEIVER_EMAILS", "").split(",") if e.strip()]

# =============================================================================
#  RA NO FILTER — Rate Agreement bids are skipped entirely
#  GeM bid numbers for RA look like: GEM/2026/R/626722  (contains /R/)
#  Regular bids look like:           GEM/2026/B/7175886 (contains /B/)
# =============================================================================
def _is_ra_bid(bid_number: str) -> bool:
    """Return True if this is a Rate Agreement bid — should be excluded."""
    # Pattern: GEM/YYYY/R/NNNNNN  — the segment after year is 'R'
    return bool(re.search(r'/R/', bid_number, re.IGNORECASE))

# =============================================================================
#  SESSION
# =============================================================================
def create_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    })

    for attempt in range(1, 4):
        try:
            session.get(BASE_URL + "/home", timeout=20)
            time.sleep(1)
            resp = session.get(ADVANCE_URL, timeout=20)
            html = resp.text

            csrf = ""
            for pat in [
                r'name=["\']csrf_bd_gem_nk["\'][^>]+value=["\']([^"\']+)["\']',
                r'value=["\']([^"\']+)["\'][^>]+name=["\']csrf_bd_gem_nk["\']',
                r'csrf_bd_gem_nk[\'"]?\s*[=:]\s*[\'"]([^\'"]+)',
                r'<meta\s+name=["\']csrf-token["\'][^>]+content=["\']([^"\']+)["\']',
                r'window\.__CSRF\s*=\s*[\'"]([^\'"]+)',
            ]:
                m = re.search(pat, html, re.I)
                if m:
                    csrf = m.group(1).strip()
                    LOG.info("  CSRF ✓")
                    break

            if not csrf:
                for cname in ("csrf_gem_cookie", "csrf_token", "CSRF-TOKEN", "_token"):
                    csrf = session.cookies.get(cname, "")
                    if csrf:
                        LOG.info(f"  CSRF from cookie ✓")
                        break

            if not csrf:
                LOG.warning(f"  Attempt {attempt}: no CSRF")
                time.sleep(2)
                continue

            session._gem_csrf = csrf
            session.headers.update({
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "X-Requested-With": "XMLHttpRequest",
                "Referer": ADVANCE_URL,
                "Origin": BASE_URL,
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            })
            LOG.info(f"  Session ready (try {attempt})")
            return session
        except Exception as e:
            LOG.warning(f"  Try {attempt} failed: {e}")
            time.sleep(3)

    LOG.error("No session after 3 tries")
    session._gem_csrf = ""
    return session

# =============================================================================
#  API FETCH
# =============================================================================
def fetch_all_bids(session: requests.Session, target: Dict) -> Tuple[List[Dict], requests.Session]:
    ministry = target["ministry"]
    org      = target.get("organization", "")
    label    = f"{ministry} / {org or 'All'}"
    LOG.info(f"  Fetching: {label}")

    all_docs: List[Dict] = []

    for page in range(1, MAX_PAGES + 1):
        csrf = getattr(session, "_gem_csrf", "")
        post_data = {
            "payload": json.dumps({
                "searchType": "ministry-search",
                "ministry": ministry,
                "buyerState": "",
                "organization": org,
                "department": "",
                "bidEndFromMin": "",
                "bidEndToMin": "",
                "from_date": "",
                "to_date": "",
                "page": page,
            }),
            "csrf_bd_gem_nk": csrf,
        }

        data = None
        for net_try in range(3):
            try:
                resp = session.post(SEARCH_URL, data=post_data, timeout=30)
            except Exception as e:
                wait = 15 * (net_try + 1)
                if net_try < 2:
                    LOG.warning(f"    Network error p{page} (try {net_try+1}/3) — wait {wait}s")
                    time.sleep(wait)
                    continue
                else:
                    LOG.error(f"    Network fail p{page} — stop")
                    return all_docs, session

            try:
                data = json.loads(resp.text)
                break
            except Exception:
                if net_try == 0:
                    LOG.warning(f"    HTTP {resp.status_code} non-JSON p{page} — refresh session")
                    time.sleep(10)
                    session = create_session()
                    post_data["csrf_bd_gem_nk"] = getattr(session, "_gem_csrf", "")
                else:
                    LOG.error(f"    Still bad — skip")
                    return all_docs, session

        if data is None:
            break

        if isinstance(data, dict) and (data.get("status") == 0 or data.get("code") == 404):
            msg = data.get("message", "no data")
            if page == 1:
                LOG.warning(f"    P1: '{msg}' — refresh & retry")
                time.sleep(DELAY_REFRESH)
                session = create_session()
                post_data["csrf_bd_gem_nk"] = getattr(session, "_gem_csrf", "")
                try:
                    resp2 = session.post(SEARCH_URL, data=post_data, timeout=30)
                    data2 = json.loads(resp2.text)
                except Exception:
                    LOG.info("    Retry fail — no bids")
                    break
                if isinstance(data2, dict) and (data2.get("status") == 0 or data2.get("code") == 404):
                    LOG.info(f"    Still no data")
                    break
                else:
                    data = data2
            else:
                LOG.info(f"    '{msg}' — done")
                break

        try:
            solr = data["response"]["response"]
            docs = solr.get("docs", [])
            num_found = int(solr.get("numFound", 0))
        except (KeyError, TypeError):
            try:
                docs = data["response"]["docs"]
                num_found = len(docs)
            except (KeyError, TypeError):
                LOG.warning(f"    Bad JSON p{page}")
                break

        if not docs:
            LOG.info(f"    P{page}: empty")
            break

        LOG.info(f"    P{page}: {len(docs)} bids")
        all_docs.extend(docs)

        if page == 1 and num_found > 0:
            pages_needed = -(-num_found // PAGE_SIZE)
            LOG.info(f"    Total: {num_found} → {pages_needed} pages")
            if pages_needed > MAX_PAGES:
                LOG.warning(f"    ⚠️  Cap hit: {num_found} bids but MAX_PAGES={MAX_PAGES}")

        if len(all_docs) >= num_found:
            LOG.info(f"    ✓ Got all {num_found}")
            break

        if page >= MAX_PAGES:
            LOG.warning(f"    ⚠️  MAX_PAGES — got {len(all_docs)}/{num_found}")
            break

        time.sleep(DELAY_PAGE)

    LOG.info(f"  ✓ Total raw: {len(all_docs)}")
    return all_docs, session

# =============================================================================
#  HELPERS
# =============================================================================
def _safe(doc: Dict, key: str, fallback="") -> str:
    val = doc.get(key, fallback)
    if isinstance(val, list):
        return str(val[0]).strip() if val else fallback
    return str(val).strip() if val else fallback

def _fmt_date(raw: str) -> str:
    if not raw:
        return ""
    for fmt in ("%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S", "%d-%m-%Y %I:%M %p", "%d-%m-%Y %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw[:19] if "T" in raw else raw, fmt).strftime("%d-%m-%Y %I:%M %p")
        except:
            pass
    return raw

def safe_filename(name: str) -> str:
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, "_")
    return name.strip(". ") or "file"

def _clean_sheet_name(name: str, max_len=31) -> str:
    for ch in r'\/:*?[]':
        name = name.replace(ch, " ")
    return re.sub(r'\s+', ' ', name).strip()[:max_len]

def get_free_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    try:
        with open(path, "a"):
            pass
        return path
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(path)
        return f"{base}_{ts}{ext}"

# =============================================================================
#  PARSE DOCS — Filters out RA bids before they enter any downstream logic
# =============================================================================
EXCEL_COLS = ["Status", "Bid Number", "Bid URL", "Category", "Quantity",
              "Ministry", "Organization", "Start Date", "End Date"]

def parse_docs(docs: List[Dict], ministry: str, organization: str) -> List[Dict]:
    bids = []
    ra_skipped = 0

    for doc in docs:
        bid_number = _safe(doc, "b_bid_number")
        if not bid_number:
            continue

        # ✅ Skip Rate Agreement (RA) bids — pattern GEM/YYYY/R/NNNNNN
        if _is_ra_bid(bid_number):
            ra_skipped += 1
            continue

        doc_id = _safe(doc, "id")
        if doc_id and doc_id.isdigit():
            bid_url = f"{BASE_URL}/showbidDocument/{doc_id}"
        else:
            bid_url = f"{BASE_URL}/biddetail/{bid_number.replace('/', '%2F')}"

        bids.append({
            "Bid Number":   bid_number,
            "Bid URL":      bid_url,
            "Category":     _safe(doc, "b_category_name"),
            "Quantity":     _safe(doc, "b_total_quantity"),
            "Ministry":     _safe(doc, "ba_official_details_minName", ministry),
            "Organization": _safe(doc, "ba_official_details_orgName", organization),
            "Start Date":   _fmt_date(_safe(doc, "final_start_date_sort")),
            "End Date":     _fmt_date(_safe(doc, "final_end_date_sort")),
        })

    if ra_skipped:
        LOG.info(f"  ⛔ RA bids filtered out: {ra_skipped}")

    return bids

# =============================================================================
#  HISTORY
# =============================================================================
def load_history() -> Dict:
    if Path(HISTORY_FILE).exists():
        try:
            with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return {}

def save_history(h: Dict):
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(h, f, indent=2, ensure_ascii=False)
        LOG.info(f"  History: {sum(len(v) for v in h.values())} bids")
    except Exception as e:
        LOG.warning(f"  History fail: {e}")

def classify_bids(bids: List[Dict], org_key: str, history: Dict) -> Tuple[List, List, List]:
    org_hist = history.get(org_key, {})
    new_l, chg_l, ex_l = [], [], []
    for bid in bids:
        bn = bid.get("Bid Number", "")
        if not bn:
            continue
        if bn not in org_hist:
            new_l.append(("NEW", bid))
        else:
            old_end = org_hist[bn].get("End Date", "")
            new_end = bid.get("End Date", "")
            if old_end and new_end and old_end != new_end:
                bid["_old_end_date"] = old_end
                chg_l.append(("DATE CHANGED", bid))
            else:
                ex_l.append(("EXISTING", bid))
    return new_l, chg_l, ex_l

def update_history(bids: List[Dict], org_key: str, history: Dict):
    if org_key not in history:
        history[org_key] = {}
    for bid in bids:
        bn = bid.get("Bid Number", "")
        if bn:
            history[org_key][bn] = {
                "End Date": bid.get("End Date", ""),
                "last_seen": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }

# =============================================================================
#  EXCEL STYLING
# =============================================================================
def _style_sheet(ws, col_index: Dict):
    url_col = col_index.get("Bid URL")
    bn_col  = col_index.get("Bid Number")

    for cell in ws[1]:
        cell.fill      = COLOR_HEADER
        cell.font      = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN
    ws.row_dimensions[1].height = 28

    for row_idx in range(2, ws.max_row + 1):
        status = (ws.cell(row_idx, col_index.get("Status", 1)).value or "").strip()

        if status == "NEW":
            fill = COLOR_NEW
        elif status == "DATE CHANGED":
            fill = COLOR_CHANGED
        elif row_idx % 2 == 0:
            fill = COLOR_ALT
        else:
            fill = COLOR_EXISTING

        for col in range(1, ws.max_column + 1):
            cell           = ws.cell(row_idx, col)
            cell.fill      = fill
            cell.font      = FONT_DATA
            cell.border    = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        if bn_col and url_col:
            url = ws.cell(row_idx, url_col).value or ""
            if str(url).startswith("http"):
                bn_c           = ws.cell(row_idx, bn_col)
                bn_c.hyperlink = str(url)
                bn_c.font      = FONT_LINK

    if url_col:
        ws.column_dimensions[get_column_letter(url_col)].hidden = True

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].hidden:
            continue
        max_len = max(
            (len(str(ws.cell(r, col).value or "")) for r in range(1, min(ws.max_row + 1, 200))),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 3, 55)

    ws.freeze_panes       = "A2"
    ws.auto_filter.ref    = ws.dimensions


def _style_duplicates_sheet(ws, col_index: Dict):
    url_col = col_index.get("Bid URL")
    bn_col  = col_index.get("Bid Number")

    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor="8B0000")
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN
    ws.row_dimensions[1].height = 28

    for row_idx in range(2, ws.max_row + 1):
        fill = COLOR_DUPE if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFCCCC")

        for col in range(1, ws.max_column + 1):
            cell           = ws.cell(row_idx, col)
            cell.fill      = fill
            cell.font      = FONT_DUPE
            cell.border    = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        if bn_col and url_col:
            url = ws.cell(row_idx, url_col).value or ""
            if str(url).startswith("http"):
                bn_c           = ws.cell(row_idx, bn_col)
                bn_c.hyperlink = str(url)
                bn_c.font      = Font(name="Arial", color="8B0000", underline="single", size=10, bold=True)

    if url_col:
        ws.column_dimensions[get_column_letter(url_col)].hidden = True

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].hidden:
            continue
        max_len = max(
            (len(str(ws.cell(r, col).value or "")) for r in range(1, min(ws.max_row + 1, 200))),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_len + 3, 55)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions


def _style_summary_sheet(ws):
    for cell in ws[1]:
        cell.fill = COLOR_HEADER
        cell.font = FONT_HEADER
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20


def _style_byorg_sheet(ws):
    for cell in ws[1]:
        cell.fill = COLOR_HEADER
        cell.font = FONT_HEADER
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12


# =============================================================================
#  EXCEL BUILDER
# =============================================================================
def build_master_excel(all_rows: List[Tuple], all_duplicates: List[Dict], file_path: str) -> str:
    file_path = get_free_path(file_path)
    os.makedirs(os.path.dirname(file_path) if os.path.dirname(file_path) else ".", exist_ok=True)

    rows = [
        [
            status,
            bid.get("Bid Number", ""),
            bid.get("Bid URL", ""),
            bid.get("Category", ""),
            bid.get("Quantity", ""),
            bid.get("Ministry", ""),
            bid.get("Organization", ""),
            bid.get("Start Date", ""),
            bid.get("End Date", ""),
        ]
        for status, bid in all_rows
    ]

    df = pd.DataFrame(rows, columns=EXCEL_COLS)

    dup_rows = [
        [
            "DUPLICATE",
            d.get("Bid Number", ""),
            d.get("Bid URL", ""),
            d.get("Category", ""),
            d.get("Quantity", ""),
            d.get("Ministry", ""),
            d.get("Organization", ""),
            d.get("Start Date", ""),
            d.get("End Date", ""),
        ]
        for d in all_duplicates
    ]
    df_dupes = pd.DataFrame(dup_rows, columns=EXCEL_COLS) if dup_rows else pd.DataFrame(columns=EXCEL_COLS)

    seen_ministries, seen_orgs = [], []
    for _, bid in all_rows:
        m = bid.get("Ministry", "").strip()
        o = bid.get("Organization", "").strip()
        if m and m not in seen_ministries:
            seen_ministries.append(m)
        if o and o not in seen_orgs:
            seen_orgs.append(o)

    LOG.info(f"  Ministries: {len(seen_ministries)} | Orgs: {len(seen_orgs)} | Duplicates: {len(all_duplicates)}")

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All_Bids", index=False)

        for ministry in seen_ministries:
            subset = df[df["Ministry"] == ministry]
            if not subset.empty:
                sn = _clean_sheet_name(ministry)
                base, n = sn, 1
                while sn in writer.sheets:
                    sn = _clean_sheet_name(f"{base} {n}")
                    n += 1
                subset.to_excel(writer, sheet_name=sn, index=False)

        for org in seen_orgs:
            subset = df[df["Organization"] == org]
            if not subset.empty:
                sn = _clean_sheet_name(org)
                base, n = sn, 1
                while sn in writer.sheets:
                    sn = _clean_sheet_name(f"{base} {n}")
                    n += 1
                subset.to_excel(writer, sheet_name=sn, index=False)

        pd.DataFrame({
            "Metric": ["Generated", "Total", "NEW", "CHANGED", "EXISTING", "DUPLICATES", "Ministries", "Orgs"],
            "Value": [
                datetime.now().strftime("%d-%m-%Y %H:%M"),
                len(df),
                len(df[df["Status"] == "NEW"]),
                len(df[df["Status"] == "DATE CHANGED"]),
                len(df[df["Status"] == "EXISTING"]),
                len(df_dupes),
                len(seen_ministries),
                len(seen_orgs),
            ],
        }).to_excel(writer, sheet_name="Summary", index=False)

        org_counts = (
            df.groupby(["Ministry", "Organization"])
            .size()
            .reset_index(name="Count")
            .sort_values(["Ministry", "Count"], ascending=[True, False])
        )
        org_counts.to_excel(writer, sheet_name="By_Org", index=False)

        if not df_dupes.empty:
            df_dupes.to_excel(writer, sheet_name="Duplicates", index=False)

    wb = load_workbook(file_path)
    for sn in wb.sheetnames:
        ws = wb[sn]
        if sn == "Summary":
            _style_summary_sheet(ws)
        elif sn == "By_Org":
            _style_byorg_sheet(ws)
        elif "Duplicates" in sn:
            col_index = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            _style_duplicates_sheet(ws, col_index)
        else:
            col_index = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            _style_sheet(ws, col_index)

    wb.save(file_path)
    LOG.info(f"  ✅ Excel: {file_path}")
    return file_path


# =============================================================================
#  EMAIL — RICH HTML FORMAT  (STARTTLS port 587 + SSL port 465 fallback)
# =============================================================================
def send_summary_email(results: List[Dict], all_rows: List[Tuple], all_changes: List[Dict],
                       all_duplicates: List[Dict], excel_path: str):
    if not (SENDER_EMAIL and APP_PASSWORD and RECEIVER_EMAILS):
        LOG.info("  Email skipped — credentials missing in .env")
        return

    total = len(all_rows)
    new_c = sum(1 for s, _ in all_rows if s == "NEW")
    chg_c = sum(1 for s, _ in all_rows if s == "DATE CHANGED")
    ex_c  = sum(1 for s, _ in all_rows if s == "EXISTING")
    dup_c = len(all_duplicates)

    run_time = datetime.now().strftime("%d-%m-%Y %I:%M %p")

    ministry_stats: Dict[str, Dict] = {}
    for r in results:
        min_name = r["label"].split(" / ")[0].strip()
        if min_name not in ministry_stats:
            ministry_stats[min_name] = {"total": 0, "new": 0, "changed": 0, "existing": 0}
        ministry_stats[min_name]["total"]    += r["total"]
        ministry_stats[min_name]["new"]      += r["new"]
        ministry_stats[min_name]["changed"]  += r["changed"]
        ministry_stats[min_name]["existing"] += r["existing"]

    parts = []
    if new_c:  parts.append(f"{new_c} NEW")
    if chg_c:  parts.append(f"{chg_c} CHANGED")
    if dup_c:  parts.append(f"{dup_c} DUPLICATES")
    subject = "🔔 GeM Bids — " + " | ".join(parts) if parts else "✅ GeM Bids — No Changes"

    ministry_rows_html = ""
    for min_name, stats in ministry_stats.items():
        ministry_rows_html += f"""
        <tr>
          <td style="padding:8px 12px; border-bottom:1px solid #e0e0e0; font-weight:600; color:#1F4E79;">{min_name}</td>
          <td style="padding:8px 12px; border-bottom:1px solid #e0e0e0; text-align:center;">{stats['total']}</td>
          <td style="padding:8px 12px; border-bottom:1px solid #e0e0e0; text-align:center; color:#2e7d32; font-weight:bold;">{stats['new']}</td>
          <td style="padding:8px 12px; border-bottom:1px solid #e0e0e0; text-align:center; color:#f57f17; font-weight:bold;">{stats['changed']}</td>
          <td style="padding:8px 12px; border-bottom:1px solid #e0e0e0; text-align:center; color:#555;">{stats['existing']}</td>
        </tr>"""

    changed_section = ""
    if all_changes:
        changed_rows = ""
        for c in all_changes[:20]:
            changed_rows += f"""
            <tr>
              <td style="padding:6px 10px; border-bottom:1px solid #ffe082; font-size:13px;">{c.get('bid_number','')}</td>
              <td style="padding:6px 10px; border-bottom:1px solid #ffe082; font-size:13px;">{c.get('description','')[:60]}</td>
              <td style="padding:6px 10px; border-bottom:1px solid #ffe082; font-size:13px; color:#c62828;">{c.get('old_date','')}</td>
              <td style="padding:6px 10px; border-bottom:1px solid #ffe082; font-size:13px; color:#2e7d32; font-weight:bold;">{c.get('new_date','')}</td>
            </tr>"""
        more_note = (
            f"<p style='font-size:12px;color:#777;'>... and {len(all_changes)-20} more. See Excel for full list.</p>"
            if len(all_changes) > 20 else ""
        )
        changed_section = f"""
        <div style="margin:24px 0;">
          <h3 style="color:#f57f17; margin-bottom:10px;">⚠️ Date Changed Tenders ({len(all_changes)})</h3>
          <table style="width:100%; border-collapse:collapse; background:#fffde7; border-radius:6px; overflow:hidden;">
            <thead>
              <tr style="background:#f9a825;">
                <th style="padding:8px 10px; text-align:left; font-size:13px;">Bid Number</th>
                <th style="padding:8px 10px; text-align:left; font-size:13px;">Description</th>
                <th style="padding:8px 10px; text-align:left; font-size:13px;">Old Date</th>
                <th style="padding:8px 10px; text-align:left; font-size:13px;">New Date</th>
              </tr>
            </thead>
            <tbody>{changed_rows}</tbody>
          </table>
          {more_note}
        </div>"""

    dup_section = ""
    if dup_c:
        dup_section = f"""
        <div style="margin:24px 0; padding:14px 18px; background:#ffebee; border-left:5px solid #c62828; border-radius:4px;">
          <strong style="color:#c62828;">&#x26D4; {dup_c} Duplicate Bid(s) Found</strong>
          <p style="margin:6px 0 0; font-size:13px; color:#555;">
            These bids appeared more than once across different org queries.
            They are listed in the <span style="color:#c62828; font-weight:bold;">Duplicates</span>
            sheet (highlighted in RED) in the attached Excel.
          </p>
        </div>"""

    html_body = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0; padding:0; background:#f4f6f9; font-family:Arial,sans-serif;">
  <div style="max-width:720px; margin:30px auto; background:#ffffff; border-radius:10px;
              box-shadow:0 2px 12px rgba(0,0,0,0.1); overflow:hidden;">

    <div style="background:linear-gradient(135deg,#1F4E79,#2e86c1); padding:28px 32px;">
      <h1 style="margin:0; color:#ffffff; font-size:22px; letter-spacing:0.5px;">
        &#x1F4CB; GeM BidPlus Tender Report
      </h1>
      <p style="margin:6px 0 0; color:#b3d4f0; font-size:14px;">Generated: {run_time}</p>
    </div>

    <div style="padding:24px 32px 0;">
      <h2 style="color:#1F4E79; margin-bottom:16px; font-size:16px; border-bottom:2px solid #e3eaf3; padding-bottom:8px;">
        Overall Summary
      </h2>
      <div style="display:flex; gap:12px; flex-wrap:wrap;">
        <div style="flex:1; min-width:110px; background:#e8f5e9; border-radius:8px; padding:16px; text-align:center;">
          <div style="font-size:28px; font-weight:bold; color:#2e7d32;">{new_c}</div>
          <div style="font-size:13px; color:#555; margin-top:4px;">NEW</div>
        </div>
        <div style="flex:1; min-width:110px; background:#fffde7; border-radius:8px; padding:16px; text-align:center;">
          <div style="font-size:28px; font-weight:bold; color:#f57f17;">{chg_c}</div>
          <div style="font-size:13px; color:#555; margin-top:4px;">DATE CHANGED</div>
        </div>
        <div style="flex:1; min-width:110px; background:#e3f2fd; border-radius:8px; padding:16px; text-align:center;">
          <div style="font-size:28px; font-weight:bold; color:#1565c0;">{ex_c}</div>
          <div style="font-size:13px; color:#555; margin-top:4px;">EXISTING</div>
        </div>
        <div style="flex:1; min-width:110px; background:#ffebee; border-radius:8px; padding:16px; text-align:center;">
          <div style="font-size:28px; font-weight:bold; color:#c62828;">{dup_c}</div>
          <div style="font-size:13px; color:#555; margin-top:4px;">DUPLICATES</div>
        </div>
        <div style="flex:1; min-width:110px; background:#f3e5f5; border-radius:8px; padding:16px; text-align:center;">
          <div style="font-size:28px; font-weight:bold; color:#6a1b9a;">{total}</div>
          <div style="font-size:13px; color:#555; margin-top:4px;">TOTAL</div>
        </div>
      </div>
    </div>

    <div style="padding:24px 32px 0;">
      <h2 style="color:#1F4E79; margin-bottom:12px; font-size:16px; border-bottom:2px solid #e3eaf3; padding-bottom:8px;">
        Ministry-wise Breakdown
      </h2>
      <table style="width:100%; border-collapse:collapse; font-size:14px;">
        <thead>
          <tr style="background:#1F4E79; color:#fff;">
            <th style="padding:10px 12px; text-align:left;">Ministry</th>
            <th style="padding:10px 12px; text-align:center;">Total</th>
            <th style="padding:10px 12px; text-align:center;">New</th>
            <th style="padding:10px 12px; text-align:center;">Changed</th>
            <th style="padding:10px 12px; text-align:center;">Existing</th>
          </tr>
        </thead>
        <tbody>{ministry_rows_html}</tbody>
        <tfoot>
          <tr style="background:#e3eaf3; font-weight:bold;">
            <td style="padding:10px 12px;">TOTAL</td>
            <td style="padding:10px 12px; text-align:center;">{total}</td>
            <td style="padding:10px 12px; text-align:center; color:#2e7d32;">{new_c}</td>
            <td style="padding:10px 12px; text-align:center; color:#f57f17;">{chg_c}</td>
            <td style="padding:10px 12px; text-align:center;">{ex_c}</td>
          </tr>
        </tfoot>
      </table>
    </div>

    <div style="padding:0 32px;">{changed_section}</div>
    <div style="padding:0 32px;">{dup_section}</div>

    <div style="padding:20px 32px;">
      <div style="background:#e8f5e9; border-left:5px solid #2e7d32; padding:14px 18px; border-radius:4px;">
        <strong style="color:#2e7d32;">Excel Attached:</strong>
        <span style="font-size:13px; color:#333; margin-left:8px;">{os.path.basename(excel_path)}</span>
        <p style="margin:6px 0 0; font-size:12px; color:#666;">
          Contains: All Bids &middot; Ministry tabs &middot; Org tabs &middot; Summary &middot; By_Org
          {"&middot; Duplicates (RED)" if dup_c else ""}
        </p>
      </div>
    </div>

    <div style="background:#f4f6f9; padding:16px 32px; text-align:center; border-top:1px solid #e0e0e0;">
      <p style="margin:0; font-size:12px; color:#999;">
        GeM BidPlus Automated Bot &nbsp;|&nbsp; {run_time} &nbsp;|&nbsp; RA bids excluded
      </p>
    </div>
  </div>
</body>
</html>"""

    plain_body = (
        f"GeM BidPlus Report | {run_time}\n\n"
        f"NEW: {new_c} | CHANGED: {chg_c} | EXISTING: {ex_c} | DUPLICATES: {dup_c} | TOTAL: {total}\n\n"
        + "\n".join(f"  {m}: T={s['total']} N={s['new']} C={s['changed']}" for m, s in ministry_stats.items())
        + f"\n\nNote: Rate Agreement (RA) bids excluded.\nExcel attached: {os.path.basename(excel_path)}"
    )

    # ── Build MIME (mixed so Excel attachment works alongside html/plain) ──
    msg_outer = MIMEMultipart("mixed")
    msg_outer["From"]    = SENDER_EMAIL
    msg_outer["To"]      = ", ".join(RECEIVER_EMAILS)
    msg_outer["Subject"] = subject

    msg_alt = MIMEMultipart("alternative")
    msg_alt.attach(MIMEText(plain_body, "plain", "utf-8"))
    msg_alt.attach(MIMEText(html_body,  "html",  "utf-8"))
    msg_outer.attach(msg_alt)

    if os.path.exists(excel_path):
        with open(excel_path, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment",
                            filename=os.path.basename(excel_path))
            msg_outer.attach(part)

    # ── Send: STARTTLS/587 first, SSL/465 fallback ──────────────────────────
    sent = False
    for method, host, port in [("STARTTLS", "smtp.gmail.com", 587),
                                ("SSL",      "smtp.gmail.com", 465)]:
        try:
            LOG.info(f"  Trying {method} port {port}...")
            if method == "STARTTLS":
                smtp = smtplib.SMTP(host, port, timeout=60)
                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()
            else:
                smtp = smtplib.SMTP_SSL(host, port, timeout=60)
                smtp.ehlo()

            smtp.login(SENDER_EMAIL, APP_PASSWORD)
            smtp.sendmail(SENDER_EMAIL, RECEIVER_EMAILS, msg_outer.as_string())
            smtp.quit()
            LOG.info(f"  ✅ Email sent via {method} → {', '.join(RECEIVER_EMAILS)}")
            sent = True
            break

        except smtplib.SMTPAuthenticationError as e:
            LOG.error(f"  ❌ Auth failed ({method}): {e}")
            LOG.error("     → Check App Password in .env (must be Gmail App Password, not account password)")
            break  # No point trying other port if credentials are wrong

        except smtplib.SMTPException as e:
            LOG.warning(f"  ⚠️  {method} SMTP error: {e} — trying next method")

        except Exception as e:
            LOG.warning(f"  ⚠️  {method} failed: {e} — trying next method")

    if not sent:
        LOG.error("  ❌ Email failed on all methods. Check firewall / App Password / 2FA status.")


# =============================================================================
#  TARGET RUNNER
# =============================================================================
def run_target(target: Dict, session: requests.Session, history: Dict) -> Tuple[Dict, requests.Session]:
    org     = target.get("organization", "") or "All"
    label   = f"{target['ministry']} / {org}"
    org_key = safe_filename(f"{target['ministry']}__{org}")

    LOG.info(f"\n{'='*65}")
    LOG.info(f"TARGET: {label}")
    LOG.info(f"{'='*65}")

    result = {
        "label": label, "status": "pending",
        "total": 0, "new": 0, "changed": 0, "existing": 0,
        "rows": [], "date_changes": [], "duplicates": [],
    }

    try:
        raw_docs, session = fetch_all_bids(session, target)

        if not raw_docs:
            LOG.info("  No bids")
            result["status"] = "no_bids"
            return result, session

        # parse_docs already filters out RA bids internally
        bids = parse_docs(raw_docs, target["ministry"], target.get("organization", ""))

        # ── Duplicate removal ─────────────────────────────────────────────
        unique_map     = {}
        duplicate_list = []
        for bid in bids:
            bn = bid.get("Bid Number", "").strip()
            if bn:
                if bn in unique_map:
                    duplicate_list.append(bid)
                else:
                    unique_map[bn] = bid
        bids = list(unique_map.values())
        LOG.info(f"  After dedup: {len(bids)} unique | {len(duplicate_list)} duplicates")
        # ─────────────────────────────────────────────────────────────────

        new_l, chg_l, ex_l = classify_bids(bids, org_key, history)
        LOG.info(f"  NEW={len(new_l)} CHG={len(chg_l)} EX={len(ex_l)}")

        date_changes = [
            {
                "bid_number":  b.get("Bid Number", ""),
                "description": b.get("Category", ""),
                "old_date":    b.get("_old_end_date", ""),
                "new_date":    b.get("End Date", ""),
            }
            for _, b in chg_l
        ]

        update_history(bids, org_key, history)

        result.update({
            "status":       "completed",
            "total":        len(bids),
            "new":          len(new_l),
            "changed":      len(chg_l),
            "existing":     len(ex_l),
            "rows":         new_l + chg_l + ex_l,
            "date_changes": date_changes,
            "duplicates":   duplicate_list,
        })

    except Exception as e:
        LOG.error(f"  Error: {e}\n{traceback.format_exc()}")
        result["status"] = "error"

    time.sleep(DELAY_ORG)
    return result, session


# =============================================================================
#  MAIN
# =============================================================================
def main():
    try:
        import ctypes
        ctypes.windll.kernel32.SetThreadExecutionState(0x80000000 | 0x00000001)
        LOG.info("✓ Sleep prevention ON")
    except:
        pass

    start = time.time()
    LOG.info("=" * 65)
    LOG.info(f"GeM Bot — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    LOG.info(f"Targets: {len(TARGETS)}  |  RA bids will be excluded automatically")
    LOG.info("=" * 65)

    os.makedirs(EXCEL_FOLDER, exist_ok=True)
    os.makedirs(LOG_FOLDER,   exist_ok=True)

    history = load_history()
    session = create_session()

    indices = [int(a) for a in sys.argv[1:] if a.isdigit()]
    targets = [TARGETS[i] for i in indices if i < len(TARGETS)] if indices else TARGETS
    LOG.info(f"Running: {len(targets)} targets")

    results, all_rows, all_changes, all_duplicates = [], [], [], []

    for i, target in enumerate(targets, 1):
        org = target.get("organization", "") or "All"
        LOG.info(f"\n>>> [{i}/{len(targets)}] {target['ministry']} / {org}")

        r, session = run_target(target, session, history)
        results.append(r)
        all_rows.extend(r.get("rows", []))
        all_changes.extend(r.get("date_changes", []))
        all_duplicates.extend(r.get("duplicates", []))

        LOG.info(f"<<< {r['status'].upper()} T={r['total']} N={r['new']} C={r['changed']} D={len(r.get('duplicates', []))}")

    save_history(history)

    if all_rows:
        master = os.path.join(EXCEL_FOLDER, f"gem_bids_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        LOG.info(f"\nBuilding Excel: {master}")
        excel = build_master_excel(all_rows, all_duplicates, master)
        LOG.info("Sending email...")
        send_summary_email(results, all_rows, all_changes, all_duplicates, excel)
    else:
        LOG.warning("⚠️  No bids found across all targets")

    LOG.info(f"\n{'='*65}")
    LOG.info("FINAL SUMMARY")
    LOG.info(f"{'='*65}")
    LOG.info(f"{'Target':<50} {'Status':<12} {'T':>6} {'N':>5} {'C':>5} {'D':>5}")
    LOG.info("-" * 65)
    for r in results:
        LOG.info(
            f"{r['label'][:49]:<50} {r['status']:<12} "
            f"{r['total']:>6} {r['new']:>5} {r['changed']:>5} {len(r.get('duplicates', [])):>5}"
        )

    total_bids    = sum(r["total"]                       for r in results)
    total_new     = sum(r["new"]                         for r in results)
    total_changed = sum(r["changed"]                     for r in results)
    total_dupes   = sum(len(r.get("duplicates", []))     for r in results)
    LOG.info("-" * 65)
    LOG.info(f"{'TOTALS':<50} {'':<12} {total_bids:>6} {total_new:>5} {total_changed:>5} {total_dupes:>5}")
    LOG.info(f"\n⏱️  Done in {time.time() - start:.1f}s")


if __name__ == "__main__":
    main()
